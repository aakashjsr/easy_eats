import openpyxl

from datetime import timedelta
from openpyxl import Workbook
from django.views.generic import TemplateView
from django.http import HttpResponseRedirect, HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum, Count

from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl.writer.excel import save_virtual_workbook

from core.permissions import SuperUserPermission
from core.constants import CANCELLED_ORDER_CHARGE_PERCENT
from restaurants.models import Restaurant
from orders.models import Order
from accounts.models import Diner


class AnalyticsOverview(TemplateView):
    """
    Renders Analytics page template
    """
    template_name = 'analytics.html'

    def get(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return HttpResponseRedirect("/admin")
        return super().get(request, *args, **kwargs)


class AnalyticsData(APIView):
    """
    Analytics Data
    """
    permission_classes = (SuperUserPermission, )
    def get(request, *args, **kwargs):
        restaurants = Restaurant.objects.all()
        today = timezone.now().date()
        last_monday = today - timedelta(days=today.weekday())
        month_first_day = today.replace(day=1)

        monthly_completed_order_earnings = float(Order.objects.filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today) & Q(status=Order.COMPLETED)).aggregate(total=Sum("total"))["total"] or 0)
        monthly_cancelled_order_earnings = float((Order.objects.filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today) & Q(status=Order.CANCELLED)).aggregate(total=Sum("total"))["total"] or 0)) * CANCELLED_ORDER_CHARGE_PERCENT
        monthly_pending_earning = float(Order.objects.filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today) & Q(status__in=[Order.ACTIVE, Order.BOOKED])).aggregate(total=Sum("total"))["total"] or 0)

        weekly_completed_order_earnings = float(
            Order.objects.filter(Q(created__date__gte=last_monday) & Q(created__date__lte=today) & Q(status=Order.COMPLETED)).aggregate(total=Sum("total"))["total"] or 0)
        weekly_cancelled_order_earnings = float((Order.objects.filter(Q(created__date__gte=last_monday) & Q(created__date__lte=today) & Q(status=Order.CANCELLED)).aggregate(
            total=Sum("total"))["total"] or 0)) * CANCELLED_ORDER_CHARGE_PERCENT
        weekly_pending_earning = float(
            Order.objects.filter(Q(created__date__gte=last_monday) & Q(created__date__lte=today) & Q(status__in=[Order.ACTIVE, Order.BOOKED])).aggregate(total=Sum("total"))[
                "total"] or 0)

        daily_completed_order_earnings = float(
            Order.objects.filter(created__date=today, status=Order.COMPLETED).aggregate(total=Sum("total"))["total"] or 0)
        daily_cancelled_order_earnings = float((Order.objects.filter(created__date=today, status=Order.CANCELLED).aggregate(
            total=Sum("total"))["total"] or 0)) * CANCELLED_ORDER_CHARGE_PERCENT
        daily_pending_earning = float(
            Order.objects.filter(created__date=today, status__in=[Order.BOOKED, Order.ACTIVE]).aggregate(total=Sum("total"))[
                "total"] or 0)

        orders_by_status = Order.objects.values("status").annotate(count=Count("status"))

        data = {
            "order_by_status": {item["status"].lower():item["count"] for item in orders_by_status},
            "repeat_orders": Order.objects.filter(is_repeat_order=True).count(),
            "sales": {
                "daily": {
                    "completed_orders": daily_completed_order_earnings,
                    "cancelled_orders": daily_cancelled_order_earnings,
                    "pending_orders": daily_pending_earning
                },
                "weekly": {
                    "completed_orders": weekly_completed_order_earnings,
                    "cancelled_orders": weekly_cancelled_order_earnings,
                    "pending_orders": weekly_pending_earning
                },
                "monthly": {
                    "completed_orders": monthly_completed_order_earnings,
                    "cancelled_orders": monthly_cancelled_order_earnings,
                    "pending_orders": monthly_pending_earning
                }
            },
            "sign_ups": {
                "daily": Diner.objects.filter(created__date=today).count(),
                "weekly": Diner.objects.filter(Q(created__date__gte=last_monday) & Q(created__date__lte=today)).count(),
                "monthly": Diner.objects.filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today)).count()
            },
            "orders": {
                "daily": Order.objects.filter(created__date=today).count(),
                "weekly": Order.objects.filter(Q(created__date__gte=last_monday) & Q(created__date__lte=today)).count(),
                "monthly": Order.objects.filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today)).count()
            },
            "restaurants": {
                "online": restaurants.filter(Q(dineout_online=True) | Q(takeaway_online=True)).count(),
                "offline": restaurants.filter(Q(dineout_online=False) & Q(takeaway_online=False)).count()
            }
        }
        return Response(data)


class RestaurantStatement(APIView):
    permission_classes = (SuperUserPermission, )

    def get(self, request, *args, **kwargs):
        headers = ["Date Ordered", "Time Ordered", "Target Date", "Target Time",
                   "Transaction id", "Order Description", "Dine-in/Pax", "Subtotal",
                   "Sales Tax", "Service Charge", "Discount", "Cancellations", "Total"]
        today = timezone.now().date()
        month_first_day = today.replace(day=1)
        restaurant = Restaurant.objects.filter(id=kwargs.get("restaurant_id")).select_related("owner").first()
        orders = Order.objects.filter(restaurant=restaurant).filter(Q(created__date__gte=month_first_day) & Q(created__date__lte=today)).prefetch_related("order_items", "food_items")

        orders_total = orders.aggregate(pre_order_total=Sum("total")).get("pre_order_total", 0)
        wb = Workbook()
        ws = wb.active
        ws.title = restaurant.name
        ws.append(["Restaurant Name", " ", " ", restaurant.name])
        ws.append(["Restaurant Address", " ", " ", restaurant.address])
        ws.append(["Contact Person", " ", " ", restaurant.owner.user.first_name])
        ws.append(["Email", " ", " ", restaurant.owner.user.email])
        ws.append([])
        ws.append(headers)
        for order in orders:
            row = [
                str(order.created.date()), str(order.created.time()), str(order.scheduled_datetime.date()), str(order.scheduled_datetime.time()),
                ""
            ]
            description = ""
            for item in order.order_items.all():
                description += "{} x {}, ".format(item.food_item.name, item.quantity)
            row.append(description)
            if order.order_type == Order.DINE_IN:
                row.append(order.seats)
            else:
                row.append("")
            row.append(order.sub_total)
            row.append(order.sub_total * (restaurant.sales_service_tax / 100))
            row.append(order.sub_total * (restaurant.service_charge_rate / 100))
            row.append(order.discount)
            row.append(order.cancellation_charge)
            row.append(order.total)
            ws.append(row)
        ws.append([])
        ws.append([""] * 11 + ["Pre order Total", orders_total])
        ws.page_setup.fitToPage = True

        # Format cells

        for row in ws.iter_rows():
            for cell in row:
                ws.cell(row=cell.row, column=cell.column).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                               wrap_text=True)
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename={}.xls'.format(restaurant.name)
        return response
